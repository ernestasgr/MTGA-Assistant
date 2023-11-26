import glob
import openpyxl
import requests
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime

page = requests.get('https://aetherhub.com/Metagame/Historic-Brawl/30/')
soup = BeautifulSoup(page.content, 'html.parser')

decks = soup.select('a.text-decoration-none')

i = 0

for deck in decks:
    deck_link = 'https://aetherhub.com' + deck.get('href')
    print(deck_link)
    deck_page = requests.get(deck_link)

    card_soup = BeautifulSoup(deck_page.content, 'html.parser')
    cards = card_soup.select('.column-wrapper.text-center')
    card_data = []
    for card in cards:
        children = card.findChildren()
        card_data.append([])
        card_data[-1].append(children[0].get('title'))
        card_data[-1].append(int(children[2].text.split(' ')[-1].split('%')[0]))
    
    deck_name = card_data[-1][0]
    card_data.sort(key=lambda data: data[1], reverse=True)
    deck_name = deck_name.replace('/', '')
    has_been_used = True
    
    if not os.path.exists(deck_name):
        os.makedirs(deck_name)
        has_been_used = False

    xlsxfiles = []
    for file in glob.glob(f"{deck_name}/*.xlsx"):
        xlsxfiles.append(file)
    if len(xlsxfiles) == 0:
        has_been_used = False

    if has_been_used:
        xlsxfiles.sort()
        previous_file = xlsxfiles[-1]
        previous_file_path = f"{os.getcwd()}\\{previous_file}"
        if os.path.exists(previous_file_path):
            previous_workbook = openpyxl.load_workbook(previous_file_path)
            previous_sheet = previous_workbook[previous_workbook.sheetnames[0]]
            previous_data = [[cell.value for cell in row] for row in previous_sheet.iter_rows()][1:]
            for row in card_data:
                for previous_row in previous_data:
                    if row[0] == previous_row[0]:
                        row.append(previous_row[2])
                        break

    # Write current data to the new xlsx file
    current_file_path = f"{deck_name}/{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Name', 'Percentage', 'In Deck'])
    for row in card_data:
        sheet.append(row)
    
    # Adjust column widths and alignment
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        for cell in column:
            cell.alignment = Alignment(wrap_text=True)
    
    workbook.save(current_file_path)

    i += 1
    print(f"{i / len(decks) * 100}%")
